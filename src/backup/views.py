import openpyxl
from django.shortcuts import render
from .models import Username, Contact

def dataload(request):
    if request.method == 'GET':
        return render(request, 'backup/data-load.html')
    if request.method == 'POST':
        user_name  = request.POST['username']
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet 1"]
        if worksheet.max_column == 2:
            excel_data = list()
            try:
                user = Username.objects.get(user_name=user_name)
            except Username.DoesNotExist:
                user = Username.objects.create(user_name=user_name)
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
                contact = Contact.objects.create(friend_name=row_data[0], contactnumber=row_data[1], username=user)
            return render(request, 'backup/data-load.html', {"excel_data":excel_data})
        return render(request, 'backup/data-load.html', {"message":"Excel Sheet is Empty/Extra Columns.."})


def dataretrieve(request):
    if request.method == 'GET':
        return render(request, 'backup/data-retrieve.html')
    if request.method == 'POST':
        search_name = request.POST['search']
        try:
            user = Username.objects.get(user_name=search_name)
            data = Contact.objects.filter(username=user)
        except Username.DoesNotExist:
            data = Contact.objects.filter(friend_name=search_name)
        return render(request, 'backup/data-retrieve.html', {"data":data, "message":"User's/Friend's Name doesn't exists."})
